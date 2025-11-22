#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
exportacion.py
---------------
Módulo de EXPORTACIÓN para Android Forensic Extractor.

Responsabilidades ÚNICAS (después del refactor):
- Copiar artefactos crudos a export/raw (cadena de custodia).
- Recibir DataFrames YA NORMALIZADOS desde procesador_legible.ForensicDataProcessor.
- Exportar esos DataFrames a:
    * CSV (export/legible)
    * Excel (resumen_forense.xlsx)
    * PDF simple (tablas)  [opcional, si reportlab está instalado]

Todo lo que sea:
    - Parseo de TXT / CSV crudos
    - Renombrar columnas
    - Ordenar filas
    - Enriquecer campos (epoch → datetime, códigos → descripciones, etc.)
debe hacerse en procesador_legible.py.

Aquí solo formateamos y guardamos, aunque podemos generar
pequeños resúmenes (agrupaciones) como parte de los reportes.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable, Optional
from dataclasses import dataclass, field

import pandas as pd

try:
    from openpyxl.utils import get_column_letter
except Exception:  # si no hay openpyxl, igual funcionará sin auto-width
    get_column_letter = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet

    _REPORTLAB_AVAILABLE = True
except Exception:
    _REPORTLAB_AVAILABLE = False


# ---------------------------------------------------------------------------
# Meta de artefactos: nombre de hoja y descripción
# ---------------------------------------------------------------------------

FORENSIC_ARTIFACT_META: dict[str, dict[str, str]] = {
    "sms": {
        "name": "SMS_MMS",
        "desc": "Mensajes SMS/MMS extraídos del dispositivo.",
    },
    "contactos": {
        "name": "CONTACTOS",
        "desc": "Contactos de agenda telefónica.",
    },
    "llamadas": {
        "name": "LLAMADAS",
        "desc": "Registro de llamadas entrantes/salientes/perdidas.",
    },
    "calendario": {
        "name": "CALENDARIO",
        "desc": "Eventos de calendario asociados a cuentas del dispositivo.",
    },
    "whatsapp_mensajes": {
        "name": "WHATSAPP_MSG",
        "desc": "Mensajes de chats de WhatsApp (normalizados por procesador_legible).",
    },
    "whatsapp_contactos": {
        "name": "WHATSAPP_CTS",
        "desc": "Contactos / chats de WhatsApp vinculados.",
    },
    "exif_media": {
        "name": "IMAGENES_EXIF",
        "desc": "Inventario de imágenes/multimedia con metadatos EXIF/GPS.",
    },
    "downloads": {
        "name": "DESCARGAS",
        "desc": "Registros del provider de descargas de Android.",
    },
    "chrome_history": {
        "name": "HIST_CHROME",
        "desc": "Historial de navegación extraído de Google Chrome.",
    },
    "gmail_mensajes": {
        "name": "GMAIL",
        "desc": "Correos electrónicos extraídos de las bases de datos de Gmail.",
    },
    "wifi_redes": {
        "name": "WIFI",
        "desc": "Redes WiFi conocidas / configuraciones de red.",
    },
    "location": {
        "name": "UBICACION",
        "desc": "Registros relacionados con ubicación (GPS / network location).",
    },
    "usagestats": {
        "name": "USAGESTATS",
        "desc": "Estadísticas de uso de aplicaciones (usagestats).",
    },
    "apks": {
        "name": "APKS",
        "desc": "APK extraídas desde el dispositivo.",
    },
    "packages": {
        "name": "PAQUETES_APPS",
        "desc": "Meta de paquetes / aplicaciones instaladas (pm/dumpsys).",
    },
}

DFMap = Dict[str, pd.DataFrame]


def get_artifact_meta(key: str) -> tuple[str, str]:
    """
    Devuelve (nombre_hoja, descripcion) amigables para un artefacto.
    """
    info = FORENSIC_ARTIFACT_META.get(
        key,
        {
            "name": key.upper(),
            "desc": f"Artefacto '{key}' exportado automáticamente.",
        },
    )
    sheet_name = info["name"][:31]  # Excel solo acepta 31 caracteres
    return sheet_name, info["desc"]


# ---------------------------------------------------------------------------
# Utils internos
# ---------------------------------------------------------------------------

def _read_text_safe(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="ignore")


def _copy_file(src: Path, dst: Path) -> None:
    if not src.exists():
        return
    dst.parent.mkdir(parents=True, exist_ok=True)
    # texto plano para preservar legibilidad en cualquier editor
    dst.write_text(_read_text_safe(src), encoding="utf-8")
    print(f"  [RAW] Copiado {src} -> {dst}")


# ---------------------------------------------------------------------------
# 1) Copia de artefactos crudos (cadena de custodia)
# ---------------------------------------------------------------------------

def copy_raw_files(
    logical_dir: Path,
    raw_dir: Path,
    extra_logical_files: Optional[Iterable[str]] = None,
) -> None:
    """
    Copia artefactos crudos relevantes a export/raw.

    logical_dir:
        - <caso>/logical
        - <caso>/noroot/logical
        - <caso>/root/logical

    Se buscan archivos en:
        logical_dir
        logical_dir.parent / "system"   (para dumpsys_location, dumpsys_wifi, etc.)
    """
    logical_dir = Path(logical_dir)
    raw_dir = Path(raw_dir)
    raw_dir.mkdir(parents=True, exist_ok=True)

    base_dir = logical_dir.parent
    sys_dir = base_dir / "system"

    # Archivos "estándar" que solemos querer preservar
    logical_names = [
        "contacts.txt",
        "calllog.txt",
        "sms.txt",
        "calendar_events.txt",
        "downloads.txt",
    ]
    if extra_logical_files:
        logical_names.extend(extra_logical_files)

    system_names = [
        "dumpsys_location.txt",
        "dumpsys_wifi.txt",
        "ip_addr.txt",
        "ip_route.txt",
        "netcfg.txt",
        "logcat_dump.txt",
        "bugreport.zip", 
    ]

    # Desde logical/
    for name in logical_names:
        src = logical_dir / name
        dst = raw_dir / name
        if src.suffix.lower() == ".zip":
            # zip → copiar binario
            if src.exists():
                dst.write_bytes(src.read_bytes())
                print(f"  [RAW] Copiado binario {src} -> {dst}")
        else:
            _copy_file(src, dst)

    # Desde system/
    for name in system_names:
        src = sys_dir / name
        dst = raw_dir / name
        if not src.exists():
            continue
        if src.suffix.lower() == ".zip":
            dst.write_bytes(src.read_bytes())
            print(f"  [RAW] Copiado binario {src} -> {dst}")
        else:
            _copy_file(src, dst)


# ---------------------------------------------------------------------------
# 2) Exportación a CSV legible
# ---------------------------------------------------------------------------

def _export_sms_csv(df: pd.DataFrame, legible_dir: Path) -> None:
    """
    CSV principal + resumen por número.
    Se asume que df viene normalizado desde procesador_legible:
    columnas: fecha_hora, numero, tipo_codigo, tipo_descripcion, mensaje
    """
    legible_dir.mkdir(parents=True, exist_ok=True)
    path_main = legible_dir / "sms_legible.csv"
    df.to_csv(path_main, index=False, encoding="utf-8-sig")
    print(f"  [CSV] {path_main.name}")

    if "numero" in df.columns:
        resumen = (
            df.groupby("numero", dropna=False)
            .agg(total_mensajes=("mensaje", "count"))
            .reset_index()
        )
        path_res = legible_dir / "sms_resumen_por_numero.csv"
        resumen.to_csv(path_res, index=False, encoding="utf-8-sig")
        print(f"  [CSV] {path_res.name}")


def _export_llamadas_csv(df: pd.DataFrame, legible_dir: Path) -> None:
    """
    CSV principal + resumen por número.
    Se asume que df viene normalizado desde procesador_legible:
    columnas: fecha_hora, numero, nombre_cache, tipo_codigo, tipo_descripcion, duracion_seg
    """
    legible_dir.mkdir(parents=True, exist_ok=True)
    path_main = legible_dir / "llamadas_legible.csv"
    df.to_csv(path_main, index=False, encoding="utf-8-sig")
    print(f"  [CSV] {path_main.name}")

    if "numero" in df.columns:
        agg_map = {
            "total_llamadas": ("tipo_codigo", "count"),
        }
        if "duracion_seg" in df.columns:
            agg_map["duracion_total_seg"] = ("duracion_seg", "sum")

        resumen = (
            df.groupby("numero", dropna=False)
            .agg(**agg_map)
            .reset_index()
        )
        path_res = legible_dir / "llamadas_resumen_por_numero.csv"
        resumen.to_csv(path_res, index=False, encoding="utf-8-sig")
        print(f"  [CSV] {path_res.name}")


def _export_generico_csv(nombre: str, df: pd.DataFrame, legible_dir: Path) -> None:
    legible_dir.mkdir(parents=True, exist_ok=True)
    path = legible_dir / f"{nombre}.csv"
    df.to_csv(path, index=False, encoding="utf-8-sig")
    print(f"  [CSV] {path.name}")


def export_csv_legible(dfs: DFMap, legible_dir: Path) -> None:
    """
    Exporta todos los artefactos disponibles en `dfs` a CSV legibles.

    IMPORTANTE:
    - Se asume que `dfs` viene de ForensicDataProcessor.load_all()
      y por tanto YA está normalizado y ordenado.
    - Aquí NO se cambia el contenido de los DataFrames, solo se generan
      archivos (y algún resumen adicional cuando tiene sentido).
    """
    legible_dir = Path(legible_dir)
    legible_dir.mkdir(parents=True, exist_ok=True)

    handled: set[str] = set()

    if "sms" in dfs:
        _export_sms_csv(dfs["sms"], legible_dir)
        handled.add("sms")

    if "llamadas" in dfs:
        _export_llamadas_csv(dfs["llamadas"], legible_dir)
        handled.add("llamadas")

    # El resto se exporta de forma genérica 1:1
    for nombre, df in sorted(dfs.items(), key=lambda x: x[0]):
        if nombre in handled:
            continue
        _export_generico_csv(nombre, df, legible_dir)


# ---------------------------------------------------------------------------
# 3) Exportación a Excel (reporte multi-hoja)
# ---------------------------------------------------------------------------

@dataclass
class ForensicExcelReport:
    """
    Construye un Excel forense con varias hojas:
    - Una hoja por artefacto (sms, llamadas, exif_media, wifi, correo, etc.).
    - Cada hoja con los datos tal y como los entregó ForensicDataProcessor.
    - Hoja RESUMEN que explica qué contiene cada hoja.

    dfs: dict[str, DataFrame] devuelto por ForensicDataProcessor.load_all()
    excel_path: ruta del .xlsx a crear.
    """
    dfs: DFMap
    excel_path: Path
    case_name: str | None = None
    device_id: str | None = None
    _sheet_info: dict[str, dict[str, str]] = field(default_factory=dict)

    # ----------------- API principal -----------------

    def build(self) -> None:
        if not self.dfs:
            print("  [XLSX] No hay DataFrames, Excel no generado.")
            return

        self.excel_path = Path(self.excel_path)
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)

        # Intento de deducir nombre de caso si no viene
        if self.case_name is None:
            try:
                # .../casos/<caso>/export/resumen_forense.xlsx
                self.case_name = self.excel_path.parent.parent.name
            except Exception:
                self.case_name = ""

        resumen_rows = []

        try:
            with pd.ExcelWriter(self.excel_path, engine="openpyxl") as writer:
                # 1) Hojas por artefacto
                for key, df in sorted(self.dfs.items(), key=lambda x: x[0]):
                    sheet_name, desc = get_artifact_meta(key)

                    # df_norm: usamos tal cual, solo garantizamos que no esté vacío
                    if df is None or df.empty:
                        df_norm = pd.DataFrame({"INFO": ["Sin registros."]})
                    else:
                        df_norm = df.copy()

                    startrow = 3  # filas 1-2 para descripción
                    df_norm.to_excel(
                        writer,
                        sheet_name=sheet_name,
                        index=False,
                        startrow=startrow,
                    )
                    ws = writer.sheets[sheet_name]

                    # Descripción en la primera fila
                    descripcion = desc or f"Artefacto '{key}' extraído del dispositivo."
                    if self.case_name:
                        descripcion = f"Caso: {self.case_name} | {descripcion}"
                    ws.cell(row=1, column=1, value=descripcion)

                    # Congelar fila de cabeceras
                    ws.freeze_panes = ws.cell(row=startrow + 1, column=1)

                    # Auto-ancho de columnas
                    if get_column_letter is not None:
                        self._auto_width(ws, df_norm)

                    # Info para hoja RESUMEN
                    resumen_rows.append(
                        {
                            "Hoja": sheet_name,
                            "Artefacto interno": key,
                            "Filas": int(df_norm.shape[0]),
                            "Columnas": int(df_norm.shape[1]),
                            "Descripción": desc,
                        }
                    )

                # 2) Hoja RESUMEN
                if resumen_rows:
                    df_resumen = pd.DataFrame(resumen_rows)
                    df_resumen.to_excel(
                        writer,
                        sheet_name="RESUMEN",
                        index=False,
                    )
                    ws_res = writer.sheets["RESUMEN"]
                    ws_res.freeze_panes = ws_res.cell(row=2, column=1)
                    if get_column_letter is not None:
                        self._auto_width(ws_res, df_resumen)

            print(f"  [XLSX] Reporte forense avanzado exportado en {self.excel_path}")
        except Exception as e:
            print(f"  [!] No se pudo crear el Excel ({self.excel_path.name}): {e}")
            print("      Aun así, los CSV legibles ya pueden estar generados.")

    # ----------------- Auto-ancho columnas -----------------

    def _auto_width(self, ws, df: pd.DataFrame) -> None:
        if get_column_letter is None:
            return
        for idx, col in enumerate(df.columns, start=1):
            try:
                serie = df[col].astype(str)
                max_len = max(
                    [len(str(col))]
                    + [len(x) for x in serie.head(200).tolist()]
                )
                max_len = min(max_len + 2, 60)
                ws.column_dimensions[get_column_letter(idx)].width = max_len
            except Exception:
                continue

    # ----------------- PDF (opcional) -----------------

    def build_pdf(self, pdf_path: Path, max_rows_per_table: int = 200) -> None:
        """
        Genera un informe PDF simple, una tabla por artefacto.
        Las columnas y el orden vienen dados por los DataFrames.
        """
        if not _REPORTLAB_AVAILABLE:
            print("  [PDF] reportlab no está instalado; se omite exportación a PDF.")
            return

        if not self.dfs:
            print("  [PDF] No hay DataFrames, PDF no generado.")
            return

        pdf_path = Path(pdf_path)
        pdf_path.parent.mkdir(parents=True, exist_ok=True)

        styles = getSampleStyleSheet()
        elements = []

        for key, df in sorted(self.dfs.items(), key=lambda x: x[0]):
            sheet_name, desc = get_artifact_meta(key)

            elements.append(Paragraph(f"Artefacto: {sheet_name} ({key})", styles["Heading2"]))
            elements.append(Paragraph(desc, styles["Normal"]))
            elements.append(Spacer(1, 6))

            if df is None or df.empty:
                df_norm = pd.DataFrame({"INFO": ["Sin registros."]})
            else:
                df_norm = df.copy()

            df_small = df_norm.head(max_rows_per_table).copy()
            data = [list(df_small.columns)] + df_small.astype(str).values.tolist()

            table = Table(data, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ]
                )
            )
            elements.append(table)
            elements.append(Spacer(1, 18))

        doc = SimpleDocTemplate(str(pdf_path), pagesize=A4)
        try:
            doc.build(elements)
            print(f"  [PDF] Resumen exportado en {pdf_path}")
        except Exception as e:
            print(f"  [!] Error generando PDF ({pdf_path.name}): {e}")


# ---------------------------------------------------------------------------
# Funciones de alto nivel
# ---------------------------------------------------------------------------

def export_excel_resumen(dfs: DFMap, excel_path: Path) -> None:
    """
    Crea un Excel con varias hojas usando ForensicExcelReport.
    Cada clave de `dfs` (sms, contactos, llamadas, calendario,
    whatsapp_mensajes, whatsapp_contactos, exif_media, packages, wifi_redes, etc.)
    se convierte en una hoja.
    """
    report = ForensicExcelReport(dfs=dfs, excel_path=excel_path)
    report.build()


def export_pdf_resumen(
    dfs: DFMap,
    pdf_path: Path,
    max_rows_per_table: int = 200,
) -> None:
    """
    Genera un PDF usando ForensicExcelReport.build_pdf().
    """
    report = ForensicExcelReport(dfs=dfs, excel_path=pdf_path)
    report.build_pdf(pdf_path=pdf_path, max_rows_per_table=max_rows_per_table)


def export_all_from_dfs(
    logical_dir: Path,
    dfs: DFMap,
    export_dir: Path,
    crear_excel: bool = True,
    crear_pdf: bool = True,
) -> None:
    """
    Flujo completo cuando YA tienes `dfs` desde ForensicDataProcessor:

        processor = ForensicDataProcessor(logical_dir)
        dfs = processor.load_all()
        export_all_from_dfs(logical_dir, dfs, case_dir / "export")

    Hace:
        - copia RAW a export/raw
        - CSV legible a export/legible
        - Excel (y PDF opcional) en export/
    """
    export_dir = Path(export_dir)
    raw_dir = export_dir / "raw"
    legible_dir = export_dir / "legible"

    print(f"\n[*] Preparando exportación en {export_dir}")
    print("[*] Copiando archivos crudos (raw)...")
    copy_raw_files(logical_dir, raw_dir)

    print("\n[*] Generando archivos legibles (CSV)...")
    export_csv_legible(dfs, legible_dir)

    if crear_excel:
        print("\n[*] Generando Excel resumen...")
        export_excel_resumen(dfs, export_dir / "resumen_forense.xlsx")

    if crear_pdf:
        print("\n[*] Generando PDF resumen...")
        export_pdf_resumen(dfs, export_dir / "resumen_forense.pdf")


def export_legible(logical_dir: Path, legible_dir: Path) -> DFMap:
    """
    Parsea los artefactos dentro de `logical_dir` usando ForensicDataProcessor
    y genera los CSV legibles en `legible_dir`.

    Devuelve el dict de DataFrames (`dfs`) para que se pueda crear el Excel.
    """
    from procesador_legible import ForensicDataProcessor

    logical_dir = Path(logical_dir)
    legible_dir = Path(legible_dir)
    legible_dir.mkdir(parents=True, exist_ok=True)

    print(f"  [LEGIBLE] Procesando {logical_dir} -> {legible_dir}")

    processor = ForensicDataProcessor(logical_dir)
    dfs = processor.load_all()   # dict[str, DataFrame]

    # Esto genera los CSV legibles (sms_legible.csv, llamadas_legible.csv, etc.)
    export_csv_legible(dfs, legible_dir)

    return dfs


def exportar_legible(
    case_root: Path,
    logical_dir: Path,
    export_dir: Optional[Path] = None,
) -> None:
    """
    Versión de alto nivel pensada para llamar desde analisis.py.

    Hace todo el flujo:
        - copia RAW a export/raw
        - genera CSV en export/legible
        - crea resumen_forense.xlsx

    Compatibilidad:
    - Si sólo se llama con (case_root, logical_dir), se asume:
        export_dir = case_root / "export"
    - Si se pasa export_dir, se usa ese.
    """
    if export_dir is None:
        export_dir = Path(case_root) / "export"
    else:
        export_dir = Path(export_dir)

    raw_dir = export_dir / "raw"
    legible_dir = export_dir / "legible"
    export_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n[*] Preparando exportación en {export_dir}")
    print("[*] Copiando archivos crudos (raw)...")
    copy_raw_files(logical_dir, raw_dir)

    print("\n[*] Generando archivos legibles (CSV)...")
    dfs = export_legible(logical_dir, legible_dir)

    print("\n[*] Generando Excel resumen...")
    export_excel_resumen(dfs, export_dir / "resumen_forense.xlsx")


def crear_resumen_excel(
    legible_dir: Path,
    nombre_archivo: str = "resumen_forense.xlsx",
) -> None:
    """
    Crea un Excel con TODAS las tablas legibles (.csv) que existan en la
    carpeta `legible_dir` (una hoja por cada CSV).

    Esta función no usa ForensicDataProcessor; solo mira los CSV que ya existen.
    Útil si el usuario borra el Excel pero conserva los CSV.
    """
    legible_dir = Path(legible_dir)
    excel_path = legible_dir.parent / nombre_archivo

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("[!] No se pudo crear el Excel: falta 'openpyxl'.")
        print("    Instálalo con: pip install openpyxl")
        return

    csv_files = sorted(legible_dir.glob("*.csv"))
    if not csv_files:
        print(f"[XLSX] No hay CSV legibles en {legible_dir}, Excel no generado.")
        return

    dfs: Dict[str, pd.DataFrame] = {}

    for csv_path in csv_files:
        try:
            df = pd.read_csv(csv_path, encoding="utf-8", errors="ignore")
        except Exception as e:
            print(f"[!] No se pudo leer {csv_path.name} como CSV ({e}), se omite.")
            continue

        if df.empty:
            continue

        sheet_name = csv_path.stem
        invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
        for ch in invalid_chars:
            sheet_name = sheet_name.replace(ch, "_")
        sheet_name = sheet_name[:31]

        dfs[sheet_name] = df

    if not dfs:
        print("[XLSX] No hay datos útiles para el Excel, no se genera archivo.")
        return

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for sheet_name, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"[XLSX] Excel resumen generado -> {excel_path}")
